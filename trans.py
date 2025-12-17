!pip install deep-translator openpyxl tqdm

from google.colab import files
from openpyxl import load_workbook
from deep_translator import GoogleTranslator, supported_languages
from tqdm import tqdm
import time
import random
import re
import concurrent.futures # 나중에 더 고도화할 때 사용할 수 있도록 import

# ==========================================
# 설정
# ==========================================
BATCH_SIZE = 50
MAX_CHAR_LIMIT = 4500 # 이 값은 deep-translator 라이브러리가 내부적으로 처리하므로, 여기서는 참조용입니다.
MAX_BATCH_RETRIES = 3 # 배치 번역 실패 시 최대 재시도 횟수
# ==========================================

# ----------------------------------------------------
# 변수 보호를 위한 마스킹/언마스킹 함수 (이전과 동일)
# ----------------------------------------------------
def mask_variables(text):
    if not isinstance(text, str):
        return text, []
    pattern = r'\{.*?\}'
    variables = re.findall(pattern, text)
    masked_text = text
    for i, var in enumerate(variables):
        placeholder = f"__VAR_{i}__"
        masked_text = masked_text.replace(var, placeholder, 1)
    return masked_text, variables

def unmask_variables(text, variables):
    if not variables:
        return text
    restored_text = text
    for i, var in enumerate(variables):
        placeholder = f"__VAR_{i}__"
        if placeholder in restored_text:
            restored_text = restored_text.replace(placeholder, var)
        else:
            pattern = f"__\s*VAR\s*_\s*{i}\s*__"
            restored_text = re.sub(pattern, var, restored_text)
    return restored_text

# ----------------------------------------------------
# 언어 코드 파싱 함수 (자동 감지용) (이전과 동일)
# ----------------------------------------------------
def parse_lang_code(cell_value):
    if not cell_value or not isinstance(cell_value, str):
        return None
    
    val = cell_value.strip()
    
    part = re.split(r'[-_]', val)[0]
    
    if len(part) == 2 and part.isalpha():
        return part.lower()
    
    return None

# ----------------------------------------------------
# 배치 번역 처리 함수 (속도 개선 로직 적용)
# ----------------------------------------------------
def process_batch_translation(translator, text_list_for_lang):
    """
    주어진 언어의 전체 텍스트 목록을 BATCH_SIZE 단위로 나누어 번역하고,
    실패 시 재시도 로직을 적용합니다.
    """
    results = []
    if not text_list_for_lang:
        return results

    # 전체 텍스트 목록을 BATCH_SIZE 단위로 처리
    for i in range(0, len(text_list_for_lang), BATCH_SIZE):
        current_sub_batch_raw = text_list_for_lang[i : i + BATCH_SIZE]

        # 1. 변수 마스킹
        masked_sub_batch = []
        sub_batch_vars = []
        for text_item in current_sub_batch_raw:
            m_text, vars_list = mask_variables(text_item)
            masked_sub_batch.append(m_text)
            sub_batch_vars.append(vars_list)

        translated_sub_batch = []
        batch_translation_successful = False

        # 2. 배치 번역 시도 (재시도 로직 포함)
        for retry_attempt in range(MAX_BATCH_RETRIES + 1):
            try:
                # 첫 시도 시에는 대기 시간 없이 바로 번역 요청
                if retry_attempt > 0:
                    # 재시도 시 지수 백오프 (2초, 4초, 8초 + 랜덤 지연)
                    sleep_time = (2 ** (retry_attempt - 1)) + random.uniform(0, 1)
                    print(f"\n   ⚠️ 배치 번역 재시도 {retry_attempt}/{MAX_BATCH_RETRIES} 중 (대기 {int(sleep_time)}초)...")
                    time.sleep(sleep_time)
                
                translated_sub_batch = translator.translate_batch(masked_sub_batch)
                batch_translation_successful = True
                break # 성공 시 재시도 루프 탈출
            except Exception as e:
                print(f"\n⚠️ 배치 번역 실패 (오류: {e}).")
                if retry_attempt == MAX_BATCH_RETRIES:
                    print("   모든 배치 재시도 실패. 개별 번역으로 폴백합니다.")
                    # 모든 배치 재시도 실패 시 개별 번역으로 폴백
                    translated_sub_batch = []
                    for idx, text_item in enumerate(masked_sub_batch):
                        try:
                            # 개별 번역 시에는 API 과부하를 줄이기 위해 작은 딜레이를 줍니다.
                            time.sleep(0.5)
                            t_text = translator.translate(text_item)
                            translated_sub_batch.append(t_text)
                        except Exception as single_e:
                            print(f"     개별 번역 실패 for text '{text_item[:min(len(text_item), 50)]}...': {single_e}")
                            translated_sub_batch.append(current_sub_batch_raw[idx]) # 원문 추가
                    batch_translation_successful = True # 폴백 성공으로 간주
                    break # 재시도 루프 탈출

        # 3. 언마스킹 및 결과 저장
        final_sub_batch = []
        if batch_translation_successful:
            for j, trans_text in enumerate(translated_sub_batch):
                if trans_text:
                    restored = unmask_variables(trans_text, sub_batch_vars[j] if j < len(sub_batch_vars) else [])
                    final_sub_batch.append(restored)
                else:
                    # 번역 실패 시 원문 유지 (또는 빈 문자열)
                    final_sub_batch.append(current_sub_batch_raw[j] if j < len(current_sub_batch_raw) else "")
        else: # 모든 재시도 및 폴백 실패 시 (거의 일어나지 않겠지만) 원문 유지
             final_sub_batch = current_sub_batch_raw

        results.extend(final_sub_batch)
    return results

# ----------------------------------------------------
# 메인 로직
# ----------------------------------------------------
print("엑셀 파일을 업로드해주세요.")
uploaded = files.upload()

if not uploaded:
    print("파일이 업로드되지 않았습니다.")
else:
    file_path = list(uploaded.keys())[0]
    print(f"업로드된 파일: {file_path}")

    try:
        wb = load_workbook(file_path)

        target_sheets = [
            "Mobile App Web", "Mobile AppWeb",
            "Admin", "Admin-에스피텍", "Admin-유비온",
            "Cms"
        ]

        # 시트 순회
        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames: continue

            print(f"\n📂 [{sheet_name}] 데이터 스캔 및 헤더 분석 중... 🚀")
            ws = wb[sheet_name]

            # 1. 헤더 행 및 소스(en-US) 위치 찾기
            header_row = None
            source_col = None
            
            for r in range(1, min(16, ws.max_row + 1)):
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).strip()
                    if val == "en-US":
                        header_row = r
                        source_col = c
                        break
                if header_row: break

            if not header_row or not source_col:
                print(f"   ⚠️ 'en-US' 컬럼을 찾을 수 없어 [{sheet_name}] 시트를 건너뜜니다.")
                continue

            # 2. 헤더 행을 분석하여 타겟 언어 컬럼들 자동 매핑
            target_cols = {} # { col_idx: 'lang_code' }
            
            for c in range(1, ws.max_column + 1):
                if c == source_col: continue

                header_val = ws.cell(row=header_row, column=c).value
                lang_code = parse_lang_code(header_val)

                if lang_code:
                    target_cols[c] = lang_code
            
            print(f"   ℹ️ 감지된 번역 대상 언어: {list(set(target_cols.values()))}")
            if not target_cols:
                print("   ⚠️ 번역할 대상 언어 컬럼(예: id-ID, ar-SA)을 찾지 못했습니다.")
                continue

            # 3. 작업 목록 생성
            tasks_by_lang = {lang: [] for lang in set(target_cols.values())}
            total_skip_count = 0
            total_add_count = 0

            for row in range(header_row + 1, ws.max_row + 1):
                en_val = ws.cell(row=row, column=source_col).value

                if en_val and str(en_val).strip():
                    en_text = str(en_val).strip()

                    for col_idx, lang_code in target_cols.items():
                        target_cell = ws.cell(row=row, column=col_idx)
                        cell_val = target_cell.value

                        if cell_val is not None and str(cell_val).strip() != "":
                            total_skip_count += 1
                            continue

                        tasks_by_lang[lang_code].append({
                            'row': row,
                            'col': col_idx,
                            'text': en_text
                        })
                        total_add_count += 1

            print(f"   ℹ️ 스캔 결과: {total_skip_count}개 셀 건너뜀, {total_add_count}개 셀 작업 예정")

            if total_add_count == 0:
                print("   ✅ 모든 작업이 이미 완료되어 있습니다.")
                continue

            # 4. 언어별 번역 또는 복사 수행
            for lang_code, tasks in tasks_by_lang.items():
                if not tasks: continue

                # 타겟 언어코드가 'en'인 경우 (en-IN, en-PH 등) -> 번역 없이 원문 복사
                if lang_code == 'en':
                    print(f"   👉 [English Variant] 영어 변형({lang_code})은 원문 복사 중... ({len(tasks)}개)")
                    for task in tasks:
                        ws.cell(row=task['row'], column=task['col']).value = task['text']
                    continue

                # 그 외 언어는 번역 진행
                print(f"   👉 [{lang_code}] 번역 진행 중... ({len(tasks)}개)")

                translator = GoogleTranslator(source='en', target=lang_code)
                texts_to_translate_for_this_lang = [t['text'] for t in tasks]
                
                # 수정된 process_batch_translation 함수 호출
                translated_texts_for_this_lang = process_batch_translation(translator, texts_to_translate_for_this_lang)

                # tqdm 진행바 표시 및 결과 엑셀에 쓰기
                with tqdm(total=len(tasks), desc=f"   Applying {lang_code} translations") as pbar:
                    for i, task in enumerate(tasks):
                        if i < len(translated_texts_for_this_lang):
                            ws.cell(row=task['row'], column=task['col']).value = translated_texts_for_this_lang[i]
                        pbar.update(1)

            print(f"   ✨ [{sheet_name}] 작업 완료")

        output_path = "NextS_AutoDetected_Faster_Updated.xlsx"
        wb.save(output_path)
        print(f"\n🎉 모든 작업 완료! 저장됨: {output_path}")
        files.download(output_path)

    except Exception as e:
        print(f"오류 발생: {e}")
        try:
            wb.save("Backup_Error.xlsx")
            files.download("Backup_Error.xlsx")
        except:
            pass
